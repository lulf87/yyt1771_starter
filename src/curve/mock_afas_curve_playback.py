"""Workbook-backed mock AFAS curve playback helpers."""

from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
import re
from typing import TYPE_CHECKING

from openpyxl import load_workbook

if TYPE_CHECKING:
    from src.application.runtime_config import RuntimeConfig

_SPACE_CHANNEL_PATTERN = re.compile(r"^Space(\d+)$", re.IGNORECASE)


@dataclass(frozen=True, slots=True)
class MockAfasCurvePlayback:
    """Immutable workbook-backed temperature/value sample series."""

    workbook_path: str
    sheet_name: str
    channel_name: str
    temperatures_celsius: tuple[float, ...]
    values: tuple[float, ...]

    @property
    def sample_count(self) -> int:
        return len(self.temperatures_celsius)

    @property
    def start_temperature_celsius(self) -> float:
        return float(self.temperatures_celsius[0])

    @property
    def end_temperature_celsius(self) -> float:
        return float(self.temperatures_celsius[-1])


def resolve_mock_afas_curve_playback(
    runtime_config: "RuntimeConfig",
    *,
    channel_name: str,
) -> MockAfasCurvePlayback | None:
    """Load the configured workbook-backed mock AFAS curve, if any."""

    workbook_path = str(runtime_config.replay.get("mock_afas_curve_path", "") or "").strip()
    if not workbook_path:
        return None

    configured_sheet_name = str(runtime_config.replay.get("mock_afas_curve_sheet", "") or "").strip() or None
    return load_mock_afas_curve_playback(
        workbook_path,
        channel_name=channel_name,
        sheet_name=configured_sheet_name,
    )


def load_mock_afas_curve_playback(
    workbook_path: str | Path,
    *,
    channel_name: str,
    sheet_name: str | None = None,
) -> MockAfasCurvePlayback:
    """Read one workbook sheet as a deterministic AFAS playback series."""

    resolved_path = _resolve_workbook_path(workbook_path)
    requested_sheet_name = sheet_name or _infer_sheet_name_for_channel(channel_name)
    return _load_cached_mock_afas_curve_playback(str(resolved_path), channel_name, requested_sheet_name)


@lru_cache(maxsize=16)
def _load_cached_mock_afas_curve_playback(
    resolved_workbook_path: str,
    channel_name: str,
    requested_sheet_name: str | None,
) -> MockAfasCurvePlayback:
    workbook_path = Path(resolved_workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Mock AFAS curve workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        selected_sheet_name = _select_sheet_name(
            workbook.sheetnames,
            channel_name=channel_name,
            requested_sheet_name=requested_sheet_name,
        )
        sheet = workbook[selected_sheet_name]

        temperatures: list[float] = []
        values: list[float] = []
        for row in sheet.iter_rows(values_only=True):
            if not row or len(row) < 2:
                continue
            try:
                temperature_celsius = float(row[0])
                value = float(row[1])
            except (TypeError, ValueError):
                continue
            temperatures.append(temperature_celsius)
            values.append(value)
    finally:
        workbook.close()

    if not temperatures:
        raise ValueError(
            "Mock AFAS curve workbook does not contain any numeric temperature/value rows: "
            f"{workbook_path}#{selected_sheet_name}"
        )

    return MockAfasCurvePlayback(
        workbook_path=str(workbook_path),
        sheet_name=selected_sheet_name,
        channel_name=channel_name,
        temperatures_celsius=tuple(temperatures),
        values=tuple(values),
    )


def _resolve_workbook_path(workbook_path: str | Path) -> Path:
    path = Path(workbook_path)
    if path.is_absolute():
        return path
    return Path(__file__).resolve().parents[2] / path


def _select_sheet_name(
    sheet_names: list[str],
    *,
    channel_name: str,
    requested_sheet_name: str | None,
) -> str:
    if not sheet_names:
        raise ValueError("Mock AFAS curve workbook does not contain any worksheets.")

    if requested_sheet_name:
        if requested_sheet_name not in sheet_names:
            raise ValueError(
                f"Mock AFAS curve workbook is missing worksheet '{requested_sheet_name}'. "
                f"Available sheets: {', '.join(sheet_names)}"
            )
        return requested_sheet_name

    inferred_sheet_name = _infer_sheet_name_for_channel(channel_name)
    if inferred_sheet_name is not None and inferred_sheet_name in sheet_names:
        return inferred_sheet_name
    if channel_name in sheet_names:
        return channel_name
    return sheet_names[0]


def _infer_sheet_name_for_channel(channel_name: str) -> str | None:
    match = _SPACE_CHANNEL_PATTERN.match(str(channel_name).strip())
    if match is None:
        return None
    return f"Sheet{int(match.group(1))}"
