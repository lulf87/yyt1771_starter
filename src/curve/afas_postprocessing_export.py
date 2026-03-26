"""PNG and Excel export helpers for the full AFAS postprocessing pipeline."""

from __future__ import annotations

import io
from typing import Any, Mapping

import matplotlib

matplotlib.use("Agg")

from matplotlib import pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def build_afas_analysis_png_bytes(
    preprocessing_result: Mapping[str, Any],
    analysis_result: Mapping[str, Any],
    *,
    channel_name: str,
) -> bytes:
    """Render the selected AFAS analysis view as a PNG image."""

    figure = _generate_afas_analysis_figure(
        preprocessing_result,
        analysis_result,
        channel_name=channel_name,
    )
    buffer = io.BytesIO()
    try:
        figure.savefig(buffer, format="png", dpi=240, bbox_inches="tight", facecolor="#fffdf9")
        buffer.seek(0)
        return buffer.read()
    finally:
        plt.close(figure)
        buffer.close()


def build_afas_excel_report_bytes(
    preprocessing_result: Mapping[str, Any],
    analysis_result: Mapping[str, Any],
    *,
    session_id: str,
    channel_name: str,
) -> bytes:
    """Export AFAS summary and processed series into an Excel workbook."""

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "analysis"
    data_sheet = workbook.create_sheet("processed_data")

    _populate_summary_sheet(summary_sheet, preprocessing_result, analysis_result, session_id=session_id, channel_name=channel_name)
    _populate_data_sheet(data_sheet, preprocessing_result, analysis_result)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _generate_afas_analysis_figure(
    preprocessing_result: Mapping[str, Any],
    analysis_result: Mapping[str, Any],
    *,
    channel_name: str,
) -> plt.Figure:
    series = dict(analysis_result.get("series", {}))
    temperatures = [float(value) for value in series.get("temperature_celsius", [])]
    values = [float(value) for value in series.get("values", [])]
    if not temperatures or not values:
        raise ValueError("analysis_result.series must include temperature_celsius and values")

    figure, axis = plt.subplots(figsize=(10, 6), dpi=240)
    axis.set_facecolor("#fffdf9")
    figure.patch.set_facecolor("#fffdf9")

    axis.plot(temperatures, values, color="#36506c", linewidth=2.8, label="Smoothed curve", zorder=3)

    fit = dict(analysis_result.get("fit", {}))
    result = dict(analysis_result.get("result", {}))
    _plot_segment(axis, fit.get("low_baseline"), "#0f766e", "Low baseline")
    _plot_segment(axis, fit.get("high_baseline"), "#b45309", "High baseline")
    _plot_tangent(axis, fit.get("tangent"), temperatures, "#cf1124", "Middle tangent")

    tangent = fit.get("tangent")
    if isinstance(tangent, dict):
        tangent_slope = float(tangent.get("slope", 0.0))
        tangent_intercept = float(tangent.get("intercept", 0.0))
        _plot_marker(axis, result.get("As"), tangent_slope, tangent_intercept, "#0f766e", "As")
        _plot_marker(axis, result.get("Af_tan"), tangent_slope, tangent_intercept, "#b45309", "Af-tan")
        _plot_marker(axis, result.get("max_slope_temp"), tangent_slope, tangent_intercept, "#cf1124", "Slope")

    axis.set_title(f"{channel_name} AFAS analysis", fontsize=14, fontweight="bold")
    axis.set_xlabel("Temperature (°C)")
    axis.set_ylabel("Metric")
    axis.grid(True, alpha=0.18)
    axis.legend(loc="best")
    return figure


def _plot_segment(axis: Any, line: Any, color: str, label: str) -> None:
    if not isinstance(line, Mapping):
        return
    range_celsius = line.get("range_celsius")
    slope = line.get("slope")
    intercept = line.get("intercept")
    if not isinstance(range_celsius, list | tuple) or len(range_celsius) != 2:
        return
    if slope is None or intercept is None:
        return
    start = float(range_celsius[0])
    end = float(range_celsius[1])
    x_values = [start, end]
    y_values = [float(slope) * start + float(intercept), float(slope) * end + float(intercept)]
    axis.plot(x_values, y_values, linestyle="--", linewidth=2.0, color=color, label=label, zorder=2)


def _plot_tangent(axis: Any, line: Any, temperatures: list[float], color: str, label: str) -> None:
    if not isinstance(line, Mapping) or not temperatures:
        return
    slope = line.get("slope")
    intercept = line.get("intercept")
    if slope is None or intercept is None:
        return
    start = min(temperatures)
    end = max(temperatures)
    x_values = [start, end]
    y_values = [float(slope) * start + float(intercept), float(slope) * end + float(intercept)]
    axis.plot(x_values, y_values, linestyle="--", linewidth=2.0, color=color, label=label, zorder=2)


def _plot_marker(axis: Any, temperature: Any, slope: float, intercept: float, color: str, label: str) -> None:
    if temperature is None:
        return
    x_value = float(temperature)
    y_value = slope * x_value + intercept
    axis.scatter([x_value], [y_value], s=70, color=color, edgecolors="#fffdf9", linewidths=1.6, zorder=4)
    axis.annotate(
        f"{label} = {x_value:.2f}°C",
        xy=(x_value, y_value),
        xytext=(10, 12),
        textcoords="offset points",
        fontsize=10,
        color=color,
        fontweight="bold",
    )


def _populate_summary_sheet(
    sheet: Any,
    preprocessing_result: Mapping[str, Any],
    analysis_result: Mapping[str, Any],
    *,
    session_id: str,
    channel_name: str,
) -> None:
    title_fill = PatternFill(start_color="6B4226", end_color="6B4226", fill_type="solid")
    title_font = Font(color="FFFFFF", bold=True, size=13)
    header_font = Font(bold=True)
    center = Alignment(vertical="center")

    sheet.merge_cells("A1:B1")
    sheet["A1"] = "AFAS Postprocessing Report"
    sheet["A1"].fill = title_fill
    sheet["A1"].font = title_font

    rows = [
        ("Session ID", session_id),
        ("Channel", channel_name),
        ("Result Status", analysis_result.get("result_status", "unavailable")),
        ("As", analysis_result.get("result", {}).get("As")),
        ("Af-tan", analysis_result.get("result", {}).get("Af_tan")),
        ("Max Slope Temp", analysis_result.get("result", {}).get("max_slope_temp")),
        ("Outlier Count", preprocessing_result.get("outlier_repair", {}).get("outlier_count", 0)),
        ("Preprocessing Warnings", "; ".join(preprocessing_result.get("warnings", []))),
        ("Analysis Warnings", "; ".join(analysis_result.get("warnings", []))),
    ]
    start_row = 3
    for index, (label, value) in enumerate(rows, start=start_row):
        sheet.cell(row=index, column=1, value=label).font = header_font
        sheet.cell(row=index, column=2, value=value if value is not None else "N/A")
        sheet.cell(row=index, column=1).alignment = center
        sheet.cell(row=index, column=2).alignment = center

    parameters = dict(preprocessing_result.get("parameters", {}))
    parameters.update(dict(analysis_result.get("parameters", {})))
    sheet.cell(row=start_row + len(rows) + 1, column=1, value="Parameter").font = header_font
    sheet.cell(row=start_row + len(rows) + 1, column=2, value="Value").font = header_font
    cursor = start_row + len(rows) + 2
    for key, value in parameters.items():
        sheet.cell(row=cursor, column=1, value=key)
        sheet.cell(row=cursor, column=2, value=str(value))
        cursor += 1

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 36


def _populate_data_sheet(
    sheet: Any,
    preprocessing_result: Mapping[str, Any],
    analysis_result: Mapping[str, Any],
) -> None:
    headers = [
        "raw_temp_c",
        "raw_value",
        "grouped_temp_c",
        "grouped_value",
        "repaired_temp_c",
        "repaired_value",
        "outlier_flag",
        "smoothed_temp_c",
        "smoothed_value",
        "derivative",
    ]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=index, value=header).font = Font(bold=True)

    raw = preprocessing_result.get("raw", {})
    grouped = preprocessing_result.get("grouped", {})
    repaired = preprocessing_result.get("outlier_repair", {})
    smoothed = preprocessing_result.get("smoothed", {})
    derivative = analysis_result.get("series", {}).get("derivative", [])
    max_length = max(
        len(raw.get("temperature_celsius", [])),
        len(grouped.get("temperature_celsius", [])),
        len(repaired.get("temperature_celsius", [])),
        len(smoothed.get("temperature_celsius", [])),
        len(derivative),
        1,
    )
    for row_index in range(max_length):
        row_number = row_index + 2
        values = [
            _value_at(raw.get("temperature_celsius", []), row_index),
            _value_at(raw.get("values", []), row_index),
            _value_at(grouped.get("temperature_celsius", []), row_index),
            _value_at(grouped.get("values", []), row_index),
            _value_at(repaired.get("temperature_celsius", []), row_index),
            _value_at(repaired.get("values", []), row_index),
            _value_at(repaired.get("outlier_mask", []), row_index),
            _value_at(smoothed.get("temperature_celsius", []), row_index),
            _value_at(smoothed.get("values", []), row_index),
            _value_at(derivative, row_index),
        ]
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_number, column=column_index, value=value)


def _value_at(values: Any, index: int) -> Any:
    if not isinstance(values, list | tuple):
        return None
    if index >= len(values):
        return None
    return values[index]
