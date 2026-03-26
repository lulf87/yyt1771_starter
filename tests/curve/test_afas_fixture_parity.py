import io
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.curve.afas_postprocessing_analysis import analyze_preprocessed_afas_channel
from src.curve.afas_postprocessing_export import (
    build_afas_analysis_png_bytes,
    build_afas_excel_report_bytes,
)
from src.curve.afas_preprocessing import preprocess_afas_channel


REPO_ROOT = Path(__file__).resolve().parents[2]
WORKSPACE_ROOT = REPO_ROOT.parent
LEGACY_AFAS_ROOT = WORKSPACE_ROOT / "AFAS"
LEGACY_REPORT_PATH = LEGACY_AFAS_ROOT / "原始数据/2026.2.26/AFReport_SP_20250507_102452.json"
LEGACY_EXPECTED_PATH = (
    REPO_ROOT / "tests/fixtures/afas/legacy_afreport_sp_20250507_102452_expected.json"
)


def _raw_rows() -> list[dict]:
    if not LEGACY_REPORT_PATH.exists():
        pytest.skip(f"Legacy AFAS representative fixture missing: {LEGACY_REPORT_PATH}")
    return json.loads(LEGACY_REPORT_PATH.read_text(encoding="utf-8"))


def _expected_fixture() -> dict:
    if not LEGACY_EXPECTED_PATH.exists():
        pytest.skip(f"Legacy AFAS parity expectation fixture missing: {LEGACY_EXPECTED_PATH}")
    return json.loads(LEGACY_EXPECTED_PATH.read_text(encoding="utf-8"))


def _build_current_dataset(channel: str, fixture: dict) -> dict:
    rows = _raw_rows()
    return {
        "schema_version": "afas_postprocessing_dataset.v1",
        "session_id": "fixture-afas-parity",
        "active_channel": channel,
        "channel_map": {
            channel: {
                "temperature_celsius": [float(row["Temperature"]) for row in rows],
                "values": [float(row[channel]) for row in rows],
                "timestamps_ms": list(range(len(rows))),
                "metric_norm": [None] * len(rows),
                "quality": [1.0] * len(rows),
                "point_a_px": [None] * len(rows),
                "point_b_px": [None] * len(rows),
            }
        },
        "preprocessing_defaults": {
            "group_by_temperature": True,
            "outlier_window": fixture["outlier_params"]["window"],
            "outlier_threshold": fixture["outlier_params"]["threshold"],
            "outlier_max_iterations": fixture["outlier_params"]["max_iterations"],
            "savgol_window_length": fixture["smooth_params"]["window_length"],
            "savgol_polyorder": fixture["smooth_params"]["polyorder"],
        },
        "analysis_defaults": {
            "low_range_celsius": list(fixture["low_range_celsius"]),
            "high_range_celsius": list(fixture["high_range_celsius"]),
            "tangent_offset": 0,
        },
    }


@pytest.mark.parametrize("channel", ("Space1", "Space3", "Space6"))
def test_representative_afreport_matches_legacy_preprocessing_and_tangent_parity(channel: str) -> None:
    fixture = _expected_fixture()
    expected = fixture["channels"][channel]
    preprocessing = preprocess_afas_channel(_build_current_dataset(channel, fixture))
    analysis = analyze_preprocessed_afas_channel(preprocessing)

    assert preprocessing["grouped"]["temperature_celsius"] == pytest.approx(expected["grouped"]["temperature_celsius"])
    assert preprocessing["grouped"]["values"] == pytest.approx(expected["grouped"]["values"])
    assert preprocessing["outlier_repair"]["temperature_celsius"] == pytest.approx(expected["outlier_repair"]["temperature_celsius"])
    assert preprocessing["outlier_repair"]["values"] == pytest.approx(expected["outlier_repair"]["values"])
    assert preprocessing["outlier_repair"]["outlier_mask"] == expected["outlier_repair"]["outlier_mask"]
    assert preprocessing["smoothed"]["temperature_celsius"] == pytest.approx(expected["smoothed"]["temperature_celsius"])
    assert preprocessing["smoothed"]["values"] == pytest.approx(expected["smoothed"]["values"])

    expected_analysis = expected["analysis"]
    assert analysis["result_status"] == "ok"
    assert analysis["result"]["As"] == pytest.approx(expected_analysis["As"])
    assert analysis["result"]["Af_tan"] == pytest.approx(expected_analysis["Af_tan"])
    assert analysis["result"]["max_slope_temp"] == pytest.approx(expected_analysis["max_slope_temp"])
    assert analysis["fit"]["low_baseline"]["slope"] == pytest.approx(expected_analysis["low_baseline"]["slope"])
    assert analysis["fit"]["low_baseline"]["intercept"] == pytest.approx(expected_analysis["low_baseline"]["intercept"])
    assert analysis["fit"]["high_baseline"]["slope"] == pytest.approx(expected_analysis["high_baseline"]["slope"])
    assert analysis["fit"]["high_baseline"]["intercept"] == pytest.approx(expected_analysis["high_baseline"]["intercept"])
    assert analysis["fit"]["tangent"]["slope"] == pytest.approx(expected_analysis["tangent"]["slope"])
    assert analysis["fit"]["tangent"]["intercept"] == pytest.approx(expected_analysis["tangent"]["intercept"])
    assert analysis["outlier_count"] == expected_analysis["outlier_count"]


def test_representative_afreport_exports_match_legacy_summary_values() -> None:
    fixture = _expected_fixture()
    channel = "Space1"
    expected_analysis = fixture["channels"][channel]["analysis"]
    preprocessing = preprocess_afas_channel(_build_current_dataset(channel, fixture))
    analysis = analyze_preprocessed_afas_channel(preprocessing)

    png_bytes = build_afas_analysis_png_bytes(preprocessing, analysis, channel_name=channel)
    workbook_bytes = build_afas_excel_report_bytes(
        preprocessing,
        analysis,
        session_id="fixture-afas-parity",
        channel_name=channel,
    )

    assert png_bytes.startswith(b"\x89PNG\r\n\x1a\n")
    assert len(png_bytes) > 10_000
    assert len(workbook_bytes) > 1_000

    workbook = load_workbook(io.BytesIO(workbook_bytes))
    assert workbook.sheetnames == ["analysis", "processed_data"]
    assert workbook["analysis"]["B6"].value == pytest.approx(expected_analysis["As"])
    assert workbook["analysis"]["B7"].value == pytest.approx(expected_analysis["Af_tan"])
    assert workbook["analysis"]["B8"].value == pytest.approx(expected_analysis["max_slope_temp"])
