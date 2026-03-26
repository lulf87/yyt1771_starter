import io

from openpyxl import load_workbook

from src.curve.afas_postprocessing_analysis import analyze_preprocessed_afas_channel
from src.curve.afas_postprocessing_export import (
    build_afas_analysis_png_bytes,
    build_afas_excel_report_bytes,
)
from src.curve.afas_preprocessing import preprocess_afas_channel


def _dataset() -> dict:
    return {
        "schema_version": "afas_postprocessing_dataset.v1",
        "session_id": "session-afas",
        "active_channel": "Space1",
        "channel_map": {
            "Space1": {
                "temperature_celsius": [25.0, 35.0, 45.0, 55.0, 65.0, 75.0, 85.0, 95.0, 105.0, 115.0],
                "values": [10.0, 10.4, 10.8, 11.6, 13.5, 18.0, 24.0, 28.4, 30.0, 30.5],
                "timestamps_ms": [1000, 1200, 1400, 1600, 1800, 2000, 2200, 2400, 2600, 2800],
                "metric_norm": [0.0, 0.02, 0.05, 0.11, 0.26, 0.55, 0.82, 0.94, 0.99, 1.0],
                "quality": [0.95] * 10,
                "point_a_px": [[20, 50]] * 10,
                "point_b_px": [[100, 50]] * 10,
            }
        },
        "preprocessing_defaults": {
            "group_by_temperature": True,
            "outlier_window": 11,
            "outlier_threshold": 5.0,
            "outlier_max_iterations": 3,
            "savgol_window_length": 5,
            "savgol_polyorder": 2,
        },
        "analysis_defaults": {
            "low_range_celsius": [25.0, 45.0],
            "high_range_celsius": [95.0, 115.0],
            "tangent_offset": 0,
        },
    }


def _analysis_pair() -> tuple[dict, dict]:
    preprocessing = preprocess_afas_channel(_dataset())
    analysis = analyze_preprocessed_afas_channel(preprocessing)
    return preprocessing, analysis


def test_build_afas_analysis_png_bytes_returns_png_signature() -> None:
    preprocessing, analysis = _analysis_pair()

    png_bytes = build_afas_analysis_png_bytes(preprocessing, analysis, channel_name="Space1")

    assert png_bytes.startswith(b"\x89PNG\r\n\x1a\n")


def test_build_afas_excel_report_bytes_returns_summary_and_processed_sheets() -> None:
    preprocessing, analysis = _analysis_pair()

    workbook_bytes = build_afas_excel_report_bytes(
        preprocessing,
        analysis,
        session_id="session-afas",
        channel_name="Space1",
    )

    workbook = load_workbook(io.BytesIO(workbook_bytes))
    assert workbook.sheetnames == ["analysis", "processed_data"]
    summary_sheet = workbook["analysis"]
    data_sheet = workbook["processed_data"]
    assert summary_sheet["A1"].value == "AFAS Postprocessing Report"
    assert summary_sheet["A3"].value == "Session ID"
    assert summary_sheet["B3"].value == "session-afas"
    assert data_sheet["A1"].value == "raw_temp_c"
    assert data_sheet["J2"].value is not None
